from openpyxl import load_workbook

# The ID and range of a sample spreadsheet.

def read_image_spreadsheet(service,  SPREADSHEET_ID, flag, RANGE_NAME = 'images'):

    if(flag==1):
        filename="Hire_Salary.xlsx"
        wb = load_workbook(filename)
        ws = wb['images']
        folder_cells,ids_cells = ws['A'] ,ws['B']
        folder = []
        ids = []
        for cell in folder_cells:
            folder.append(cell.value)
            
        for cell in ids_cells:
            ids.append(cell.value)

        return dict(zip(folder[1:],ids[1:]))
    else:
        # Call the Sheets API
        sheet = service.spreadsheets()
        result = sheet.values().get(spreadsheetId=SPREADSHEET_ID,
                                    range=RANGE_NAME).execute()
        values = result.get('values', [])
        retVal = {}
        if not values:
            print('No data found.')
        else:
            for row in values:
                # Print columns A and E, which correspond to indices 0 and 4.
                retVal[row[0]] = row[1]
            return retVal

def read_item_spreadsheet(service, SPREADSHEET_ID, flag,RANGE_NAME = 'item assignment!B2:B'):
    if(flag==1):
        filename = "Hire_Salary.xlsx"
        wb = load_workbook(filename)
        ws = wb['item assignment']
        images_cell,cond_cell =  ws['B'], ws['BV']
        images = []
        cond = []
        for cell in images_cell:
            images.append(cell.value)
            
        for cell in cond_cell:
            cond.append(cell.value)

        return images[1:],cond[1:]
    else:
        sheet = service.spreadsheets()
        range_names = [
            'item assignment!B2:B',
            'item assignment!BV2:BV'
        ]
        result = sheet.values().batchGet(spreadsheetId = SPREADSHEET_ID, ranges = range_names).execute()
        values = result.get('valueRanges', [])
        print(values)
        return values[0]['values'], values[1]['values']
