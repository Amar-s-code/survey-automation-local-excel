# The ID and range of a sample spreadsheet.
from openpyxl import load_workbook
RANGE_NAME = 'images'



def write_to_spreadsheet(values, service, SPREADSHEET_ID,flag):

    if(flag==1):

        #Write to local excel spreadsheet 
        filepath = DATASHEET
        wb=load_workbook(filepath)
        ws = wb.create_sheet(RANGE_NAME)
        for row in values:
            ws.append(row)
        wb.save(DATASHEET)

    else:
        body = {
            'values':values
        }
        # Call the Sheets API
        sheet = service.spreadsheets().values().update(
            spreadsheetId = SPREADSHEET_ID, range = RANGE_NAME, 
            valueInputOption = 'RAW', body = body).execute()

        print('{0} cells updated.'.format(sheet.get('updatedCells')))
        
